"""
MCC Research Explorer — Flask Backend
======================================
Endpoints:
  GET /                         → main page
  GET /search?q=keyword         → keyword search across cached publications
  GET /search_name?name=...     → researcher lookup (ORCID + PI_ID + publications)
  GET /search_funding?pi_id=... → NIH Reporter funding by PI_ID
  GET /search_funding?name=...  → NIH Reporter funding by researcher name
  GET /researcher?name=...      → combined: publications + funding
  GET /api/members              → list all members (for autocomplete)

Data sources:
  - Local SQLite (capstone.db): cached publications + NIH projects
  - PubMed E-Utilities (live, by ORCID)
  - NIH Reporter API (live, by PI_ID)
  - RePORTER_PI_IDS_FY2025.xlsx Sheet2: member roster (name, PI_ID, ORCID)
"""

import os
import sys
import sqlite3
import json
import time
import urllib.request
import urllib.parse
import ssl
import xml.etree.ElementTree as ET

import pandas as pd
import openpyxl
import requests
from flask import Flask, jsonify, render_template, request

# Allow `lib/` modules to be imported
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "lib"))

# ─── Paths ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "capstone.db")
PI_EXCEL = os.path.join(BASE_DIR, "data", "RePORTER_PI_IDS_FY2025.xlsx")
MEMBERS_EXCEL = os.path.join(BASE_DIR, "data", "MCC_Members_Jan_2026.xlsx")
VIVO_EXCEL = os.path.join(BASE_DIR, "data", "All_MCC_Members_FY25.xlsx")

# Tell capstone4 (and any descendant) which DB to use, before importing
os.environ.setdefault("CAPSTONE_DB", DB_PATH)

from capstone4 import lookup_publications  # noqa: E402

# ─── Constants ──────────────────────────────────────────────────────
NIH_API = "https://api.reporter.nih.gov/v2/projects/search"
SSL_CTX = ssl.create_default_context()

app = Flask(__name__, static_folder="static", template_folder="templates")


# ════════════════════════════════════════════════════════════════════
# Member roster (from Excel Sheet2)
# ════════════════════════════════════════════════════════════════════
_MEMBERS_CACHE = None

PROGRAM_MAP = {
    "Meyer Cancer Center: CB": "CB",
    "Meyer Cancer Center: CGE": "CGE",
    "Meyer Cancer Center: CPC": "CPC",
    "Meyer Cancer Center: CT": "CT",
    "Meyer Cancer Center: ZY": "ZY",
    "Meyer Cancer Center": "MCC",
}


def _load_program_data():
    """Load program affiliations from MCC Members and VIVO links from the FY25 roster."""
    program_map = {}
    vivo_map = {}

    # Program affiliations + email-based VIVO fallback from MCC Members Jan 2026
    try:
        mdf = pd.read_excel(MEMBERS_EXCEL)
        mdf = mdf.fillna("")
        for _, row in mdf.iterrows():
            name = str(row.get("Name", "")).strip()
            if not name:
                continue
            prog_raw = str(row.get("Center: Program Area", "")).strip()
            program_map[name.lower()] = PROGRAM_MAP.get(prog_raw, prog_raw)
            email = str(row.get("Contact: Email", "")).strip()
            if email and "@" in email:
                cwid = email.split("@")[0]
                vivo_map.setdefault(name.lower(),
                                    f"https://vivo.weill.cornell.edu/display/cwid-{cwid}")
    except Exception as e:
        print(f"[WARN] Could not load MCC members file: {e}")

    # VIVO links from FY25 roster (uses CWID column + hyperlinks)
    try:
        wb = openpyxl.load_workbook(VIVO_EXCEL, data_only=True)

        # Research Program members sheet has CWID + Name + Program
        ws = wb["Research Program members"]
        for row_cells in ws.iter_rows(min_row=2, max_col=3):
            cwid_cell, name_cell, prog_cell = row_cells[0], row_cells[1], row_cells[2]
            name = str(name_cell.value or "").strip()
            cwid = str(cwid_cell.value or "").strip()
            if name and cwid:
                vivo_map[name.lower()] = f"https://vivo.weill.cornell.edu/display/cwid-{cwid}"

        # ZY-clinical members sheet (Name column only, some have hyperlinks)
        ws2 = wb["ZY-clinical members"]
        for row_cells in ws2.iter_rows(min_row=2, max_col=1):
            cell = row_cells[0]
            name = str(cell.value or "").strip()
            if not name:
                continue
            if cell.hyperlink and cell.hyperlink.target and "vivo.weill.cornell.edu" in cell.hyperlink.target:
                vivo_map[name.lower()] = cell.hyperlink.target

        wb.close()
    except Exception as e:
        print(f"[WARN] Could not load VIVO roster: {e}")

    return program_map, vivo_map


_PROGRAM_CACHE = None
_VIVO_CACHE = None


def _get_program_and_vivo():
    global _PROGRAM_CACHE, _VIVO_CACHE
    if _PROGRAM_CACHE is None:
        _PROGRAM_CACHE, _VIVO_CACHE = _load_program_data()
    return _PROGRAM_CACHE, _VIVO_CACHE


def _fuzzy_lookup(data_map, name):
    """Match by exact key first, then by last name + first-name prefix."""
    key = name.lower()
    if key in data_map:
        return data_map[key]

    if "," not in key:
        return ""
    last, first = key.split(",", 1)
    last = last.strip()
    first = first.strip().split()[0] if first.strip() else ""

    for map_key, val in data_map.items():
        if "," not in map_key:
            continue
        m_last, m_first = map_key.split(",", 1)
        m_last = m_last.strip()
        m_first = m_first.strip().split()[0] if m_first.strip() else ""
        if last == m_last and first and m_first and (
            first.startswith(m_first) or m_first.startswith(first)
        ):
            return val

    return ""


def load_members():
    """Load all MCC members from Sheet2 of the Excel file."""
    global _MEMBERS_CACHE
    if _MEMBERS_CACHE is not None:
        return _MEMBERS_CACHE

    df = pd.read_excel(PI_EXCEL, sheet_name="Sheet2")
    df = df.fillna("")
    program_data, vivo_data = _get_program_and_vivo()
    members = []
    for _, row in df.iterrows():
        name = str(row.get("PI_NAMEs", "")).strip()
        if not name or name.lower() == "nan":
            continue
        pi_id = str(row.get("PI_IDS", "")).strip()
        orcid = str(row.get("ORCID", "")).strip()
        pub_count = row.get("PUB_COUNT", "")
        program = _fuzzy_lookup(program_data, name)
        vivo_url = _fuzzy_lookup(vivo_data, name)
        members.append({
            "name": name,
            "pi_id": "" if pi_id.lower() == "nan" else pi_id,
            "orcid": "" if orcid.lower() == "nan" else orcid,
            "pub_count": pub_count if pub_count != "" else None,
            "program": program,
            "vivo_url": vivo_url,
        })
    _MEMBERS_CACHE = members
    return members


# ─── MCC author matching (server-side, mirrors frontend logic) ────
_MCC_LAST_NAME_IDX = None


def _build_mcc_name_index():
    """Build a last-name → [first-name-parts] index for author matching."""
    global _MCC_LAST_NAME_IDX
    if _MCC_LAST_NAME_IDX is not None:
        return _MCC_LAST_NAME_IDX

    idx = {}
    for m in load_members():
        raw = m["name"].lower().strip()
        if "," in raw:
            last, firsts = raw.split(",", 1)
            last = last.strip()
            firsts = firsts.strip()
        else:
            parts = raw.split()
            if len(parts) >= 2:
                last = parts[-1]
                firsts = " ".join(parts[:-1])
            else:
                last = raw
                firsts = ""
        if not last:
            continue
        idx.setdefault(last, []).append(firsts)
        if "-" in last:
            for part in last.split("-"):
                if len(part) >= 2:
                    idx.setdefault(part, []).append(firsts)

    _MCC_LAST_NAME_IDX = idx
    return idx


def _first_name_matches(author_first: str, member_first: str) -> bool:
    """Strict first-name check to avoid false positives like Allison/Aaron.

    - If author first name is 1 char (initial only), accept initial match.
    - If author first name is 2+ chars, require first 2 chars to match.
    """
    if not author_first or not member_first:
        return False
    if len(author_first) == 1:
        return author_first[0] == member_first[0]
    return (author_first[:2] == member_first[:2]
            or author_first.startswith(member_first)
            or member_first.startswith(author_first))


def _is_mcc_author(author_str: str) -> bool:
    """Check if an author name matches any MCC member (fuzzy but strict on first name)."""
    idx = _build_mcc_name_index()
    author = author_str.lower().strip()
    parts = author.split()
    if len(parts) < 2:
        return False

    for n in range(1, min(4, len(parts))):
        last = " ".join(parts[len(parts) - n:])
        first_parts = parts[:len(parts) - n]
        first = first_parts[0] if first_parts else ""

        candidates = idx.get(last)
        if candidates:
            for c_firsts in candidates:
                if not c_firsts:
                    return True
                c_first = c_firsts.split()[0]
                if _first_name_matches(first, c_first):
                    return True

        if n > 1:
            hyphenated = "-".join(parts[len(parts) - n:])
            candidates = idx.get(hyphenated)
            if candidates:
                for c_firsts in candidates:
                    if not c_firsts or not first:
                        return True
                    c_first = c_firsts.split()[0]
                    if _first_name_matches(first, c_first):
                        return True
    return False


def _pub_has_mcc_author(pub: dict) -> bool:
    """Return True if at least one author in the publication is an MCC member.

    Uses multiple signals in priority order:
      1. ORCID match (author's embedded ORCID matches an MCC member's)
      2. Name match + Weill Cornell affiliation
      3. Name match alone (for papers without affiliation data)
    """
    # Build ORCID lookup for MCC members
    mcc_orcids = {m["orcid"] for m in load_members() if m["orcid"]}

    author_details = pub.get("author_details", [])
    if author_details:
        for ad in author_details:
            # Signal 1: ORCID match — strongest evidence
            if ad.get("orcid") and ad["orcid"] in mcc_orcids:
                return True
            # Signal 2: Name match + WCM affiliation
            if ad.get("name") and _is_mcc_author(ad["name"]):
                aff = (ad.get("affiliation") or "").lower()
                if _is_wcm_affiliation(aff):
                    return True
        # Signal 3: Name match without affiliation data (fallback)
        for ad in author_details:
            if ad.get("name") and _is_mcc_author(ad["name"]):
                aff = (ad.get("affiliation") or "").lower()
                if not aff:
                    # No affiliation info — accept name match
                    return True
        return False

    # Fallback: no author_details, use the old string-based check
    authors_str = pub.get("authors", "")
    if not authors_str:
        return False
    for author in authors_str.split(";"):
        author = author.strip()
        if author and _is_mcc_author(author):
            return True
    return False


# Common patterns for Weill Cornell Medicine affiliations
_WCM_AFF_PATTERNS = [
    "weill cornell", "cornell medical", "meyer cancer",
    "memorial sloan", "mskcc", "new york presbyterian",
    "nyp", "newyork-presbyterian",
]


def _is_wcm_affiliation(aff_lower: str) -> bool:
    """Check if an affiliation string indicates WCM/MCC/MSK."""
    if not aff_lower:
        return False
    return any(pat in aff_lower for pat in _WCM_AFF_PATTERNS)


def _get_orcid_confirmed_pmids(orcid: str) -> set:
    """Query PubMed for PMIDs linked to a specific ORCID."""
    if not orcid:
        return set()
    try:
        r = requests.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi",
            params={"db": "pubmed", "term": f"{orcid}[auid]",
                    "retmax": 500, "retmode": "json"},
            timeout=15)
        r.raise_for_status()
        return set(r.json().get("esearchresult", {}).get("idlist", []))
    except Exception:
        return set()


def _pub_matches_researcher(pub: dict, researcher: str,
                            member: dict = None) -> bool:
    """Check if a publication is genuinely by the specified researcher.

    When the MCC member is known (resolved from roster):
      1. ORCID-confirmed PMID → accept
      2. Author ORCID in XML matches member's ORCID → accept
      3. Name match + WCM affiliation → accept
      4. Name match + no affiliation data → accept (benefit of doubt)
      5. Name match + non-WCM affiliation → REJECT (likely different person)

    When no member is resolved, just check if the search text appears
    in any author name (for middle-name searches like "todd").
    """
    r_lower = researcher.lower().strip()

    if not member:
        # No specific member resolved (single-word search like "todd").
        # Check if any author with a matching name has WCM affiliation
        # or at least no conflicting affiliation.
        author_details = pub.get("author_details", [])
        if author_details:
            for ad in author_details:
                a_name = (ad.get("name") or "").lower()
                if r_lower not in a_name:
                    continue
                # Name matches — also must be MCC author
                if not _is_mcc_author(ad["name"]):
                    continue
                a_aff = (ad.get("affiliation") or "").lower()
                # WCM affiliation or no affiliation: accept
                if not a_aff or _is_wcm_affiliation(a_aff):
                    return True
                # Non-WCM affiliation: reject this author
            return False
        # No author_details — fall back to substring
        return r_lower in pub.get("authors", "").lower()

    # We have a specific MCC member
    member_orcid = member.get("orcid", "")
    member_name = member.get("name", "")

    # Parse member name parts for matching
    if "," in member_name:
        m_last, m_first_full = [s.strip() for s in member_name.split(",", 1)]
        m_first = m_first_full.split()[0] if m_first_full else ""
    else:
        parts = member_name.split()
        m_last = parts[-1] if parts else ""
        m_first = parts[0] if len(parts) > 1 else ""
    m_last_lower = m_last.lower()
    m_first_lower = m_first.lower()

    author_details = pub.get("author_details", [])
    if not author_details:
        # No detailed data — fall back to substring match
        return r_lower in pub.get("authors", "").lower()

    for ad in author_details:
        a_name = (ad.get("name") or "").lower()
        a_orcid = ad.get("orcid") or ""
        a_aff = (ad.get("affiliation") or "").lower()

        # Check if this author could be the member
        # First: does the name match?
        name_parts = a_name.split()
        if len(name_parts) < 2:
            continue
        a_last = name_parts[-1]
        a_first = name_parts[0]

        if a_last != m_last_lower:
            continue
        if not _first_name_matches(a_first, m_first_lower):
            continue

        # Name matches. Now verify:
        # 1. ORCID match in XML → confirmed
        if member_orcid and a_orcid == member_orcid:
            return True

        # 2. WCM affiliation → very likely the right person
        if _is_wcm_affiliation(a_aff):
            return True

        # 3. No affiliation data → accept (can't disprove)
        if not a_aff:
            return True

        # 4. Non-WCM affiliation → likely different person, skip

    return False


def find_member_by_name(query: str):
    """Case-insensitive partial name match. Returns first match or None."""
    results = find_members_fuzzy(query, limit=1)
    return results[0]["member"] if results else None


def find_members_fuzzy(query: str, limit: int = 20):
    """Return all members matching the query with match-type labels.

    Returns list of {"member": {...}, "match_type": "exact"|"partial"|"fuzzy"}.
    """
    if not query:
        return []
    q = query.strip().lower()
    members = load_members()
    seen = set()
    results = []

    def _add(m, mtype):
        key = m["name"].lower()
        if key not in seen:
            seen.add(key)
            results.append({"member": m, "match_type": mtype})

    # Exact match
    for m in members:
        if m["name"].lower() == q:
            _add(m, "exact")

    # Partial / substring match
    for m in members:
        if q in m["name"].lower():
            _add(m, "partial")

    # Fuzzy: last-name match, first-initial match, or query is a prefix
    q_parts = q.split(",") if "," in q else q.split()
    q_last = q_parts[0].strip() if q_parts else q
    q_first = q_parts[1].strip().split()[0] if len(q_parts) > 1 and q_parts[1].strip() else ""

    for m in members:
        name_lower = m["name"].lower()
        parts = name_lower.split(",")
        m_last = parts[0].strip() if parts else ""
        m_first = parts[1].strip().split()[0] if len(parts) > 1 and parts[1].strip() else ""

        # Last name contains query or query contains last name
        if q_last and (q_last in m_last or m_last.startswith(q_last)):
            if q_first:
                if m_first and (m_first.startswith(q_first) or q_first.startswith(m_first)
                                or m_first[0] == q_first[0]):
                    _add(m, "fuzzy")
            else:
                _add(m, "fuzzy")

    return results[:limit]


def find_members_by_query(query: str, limit: int = 20):
    """Return all members matching the query (for autocomplete API)."""
    results = find_members_fuzzy(query, limit=limit)
    return [r["member"] for r in results]


def find_member_by_pi_id(pi_id):
    """Reverse lookup: given a PI_ID, return the matching member (or None)."""
    if pi_id in (None, ""):
        return None
    target = str(pi_id).strip().rstrip(".0").rstrip(".")
    for m in load_members():
        mid = str(m.get("pi_id") or "").strip().rstrip(".0").rstrip(".")
        if mid and mid == target:
            return m
    return None


# ════════════════════════════════════════════════════════════════════
# Database helpers
# ════════════════════════════════════════════════════════════════════
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ════════════════════════════════════════════════════════════════════
# Click tracking helpers
# ════════════════════════════════════════════════════════════════════
CLICK_DB_PATH = os.path.join(BASE_DIR, "data", "clicks.db")


def init_click_db():
    """Create click tracking database if it does not exist."""
    conn = sqlite3.connect(CLICK_DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS click_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            element TEXT,
            page TEXT,
            timestamp TEXT,
            user_agent TEXT
        )
    """)
    conn.commit()
    conn.close()


init_click_db()

# ════════════════════════════════════════════════════════════════════
# PubMed lookup helpers
# ════════════════════════════════════════════════════════════════════
def pubmed_search_by_name(name: str, retmax: int = 30):
    """Live search PubMed by researcher name. Returns list of articles."""
    # Convert "LAST, FIRST" → "First Last"
    if "," in name:
        last, first = [s.strip() for s in name.split(",", 1)]
        first = first.split()[0] if first else ""
        query = f'{last} {first}[Author]'
    else:
        query = f'{name}[Author]'

    esearch = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {"db": "pubmed", "term": query, "retmax": retmax, "retmode": "json"}
    r = requests.get(esearch, params=params, timeout=30)
    r.raise_for_status()
    pmids = r.json().get("esearchresult", {}).get("idlist", [])

    if not pmids:
        return []

    efetch = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    r2 = requests.get(efetch, params={"db": "pubmed", "id": ",".join(pmids), "retmode": "xml"}, timeout=60)
    r2.raise_for_status()
    return _parse_pubmed_xml(r2.text)


_NIH_SUFFIXES = [
    "", "s", "er", "ers", "ing", "ed", "tion", "sion", "ment", "ive",
    "al", "ic", "ics", "ous", "ary", "ance", "ence", "ity", "ies",
    "ogy", "oma", "omas", "emia", "itis", "osis", "ases",
    "therapy", "ology", "ological",
]


def _expand_query_for_nih(text: str) -> str:
    """Expand partial terms for NIH Reporter by adding common suffixes.

    NIH Reporter wildcards don't work with pi_profile_ids, so we generate
    plausible completions and use the 'or' operator. Non-matching expansions
    are harmless.
    """
    words = text.strip().split()
    expanded = set()
    for w in words:
        expanded.add(w)
        if len(w) >= 3 and not w[-1].isdigit() and w.upper() not in ("AND", "OR", "NOT"):
            for suf in _NIH_SUFFIXES:
                expanded.add(w + suf)
    return " ".join(expanded)


def _add_wildcards(text: str, join_op: str = " ") -> str:
    """Append truncation wildcard (*) to terms ≥3 chars."""
    words = text.strip().split()
    out = []
    for w in words:
        if w.endswith("*") or w.endswith("]") or w.upper() in ("AND", "OR", "NOT"):
            out.append(w)
        elif len(w) >= 3 and not w[-1].isdigit():
            out.append(w + "*")
        else:
            out.append(w)
    return join_op.join(out)


_MCC_AUTHOR_TERMS_CACHE = None


def _get_mcc_author_terms():
    """Build PubMed author terms: ORCIDs as [auid] + names as [Author]."""
    global _MCC_AUTHOR_TERMS_CACHE
    if _MCC_AUTHOR_TERMS_CACHE is not None:
        return _MCC_AUTHOR_TERMS_CACHE

    terms = []
    for m in load_members():
        # ORCID-based (most precise)
        if m["orcid"]:
            terms.append(f'{m["orcid"]}[auid]')
        # Name-based (catches publications without ORCID tags)
        name = m["name"]
        if "," in name:
            last, first = [s.strip() for s in name.split(",", 1)]
            first = first.split()[0] if first else ""
            if first:
                terms.append(f"{last} {first}[Author]")

    _MCC_AUTHOR_TERMS_CACHE = terms
    return terms


def pubmed_search_by_keyword(keyword: str, year_start: str = "",
                             year_end: str = "", retmax: int = 50):
    """Live PubMed keyword search scoped to MCC members via ORCIDs + names.

    Uses both ORCID [auid] and member names [Author] in one query to ensure
    comprehensive coverage: ORCIDs provide precision, names catch publications
    without ORCID tags. POST handles the large query (~16KB).
    """
    author_terms = _get_mcc_author_terms()

    kw_exact = keyword.strip()
    # If user quoted an exact phrase, skip fuzzy expansion — PubMed
    # handles quoted phrases natively as exact matches.
    if kw_exact.startswith('"') and kw_exact.endswith('"'):
        kw_part = kw_exact
    else:
        kw_fuzzy = _add_wildcards(kw_exact)
        kw_part = f"({kw_exact} OR {kw_fuzzy})"

    member_part = " OR ".join(author_terms)
    term = f"{kw_part} AND ({member_part})"

    if year_start or year_end:
        mindate = year_start or "2000"
        maxdate = year_end or "2026"
        term += f" AND {mindate}[PDAT]:{maxdate}[PDAT]"

    esearch = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    # Use POST to handle large member list in query
    r = requests.post(esearch, data={
        "db": "pubmed", "term": term, "retmax": retmax,
        "retmode": "json", "sort": "date",
    }, timeout=60)
    r.raise_for_status()
    data = r.json().get("esearchresult", {})
    pmids = data.get("idlist", [])
    total = int(data.get("count", 0))

    if not pmids:
        return [], total

    efetch = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    r2 = requests.get(efetch, params={"db": "pubmed", "id": ",".join(pmids),
                                       "retmode": "xml"}, timeout=60)
    r2.raise_for_status()
    return _parse_pubmed_xml(r2.text), total


_MONTHS = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def _extract_date(article_el):
    """Return (display_date, iso_date) for sorting + UI display."""
    import re

    def _txt(parent, tag):
        if parent is None:
            return ""
        el = parent.find(tag)
        return (el.text or "").strip() if el is not None else ""

    # Prefer ArticleDate (electronic pub date)
    date_el = article_el.find(".//ArticleDate")
    if date_el is not None:
        y, m, d = _txt(date_el, "Year"), _txt(date_el, "Month"), _txt(date_el, "Day")
        if y:
            mm = _MONTHS.get(m, m).zfill(2) if m else "01"
            dd = d.zfill(2) if d else "01"
            return (f"{y}-{mm}-{dd}", f"{y}-{mm}-{dd}")

    # Fall back to PubDate
    pd_el = article_el.find(".//PubDate")
    if pd_el is not None:
        y = _txt(pd_el, "Year")
        m = _txt(pd_el, "Month")
        medline = _txt(pd_el, "MedlineDate")
        if y:
            mm = _MONTHS.get(m, m).zfill(2) if m else "01"
            return (f"{y}{(' ' + m) if m else ''}", f"{y}-{mm}-01")
        if medline:
            yr = re.match(r"\d{4}", medline)
            return (medline, f"{yr.group(0)}-01-01" if yr else "")

    return ("", "")


def _parse_pubmed_xml(xml_text: str):
    root = ET.fromstring(xml_text)
    out = []
    for art in root.findall(".//PubmedArticle"):
        pmid = art.findtext(".//PMID", default="")
        title_el = art.find(".//ArticleTitle")
        title = "".join(title_el.itertext()) if title_el is not None else ""
        abstract_parts = art.findall(".//Abstract/AbstractText")
        abstract = " ".join("".join(a.itertext()) for a in abstract_parts)
        journal = art.findtext(".//Journal/Title", default="")
        pub_date, iso_date = _extract_date(art)
        authors = []
        author_details = []  # [{name, orcid, affiliation}, ...]
        for au in art.findall(".//Author"):
            ln = au.findtext("LastName", default="")
            fn = au.findtext("ForeName", default="")
            if ln:
                full = f"{fn} {ln}".strip()
                authors.append(full)
                orcid_el = au.find(".//Identifier[@Source='ORCID']")
                orcid_val = ""
                if orcid_el is not None and orcid_el.text:
                    # Normalize: strip URL prefix if present
                    orcid_val = orcid_el.text.strip().replace(
                        "http://orcid.org/", "").replace(
                        "https://orcid.org/", "")
                affs = [a.text for a in au.findall(
                    ".//AffiliationInfo/Affiliation") if a.text]
                author_details.append({
                    "name": full,
                    "orcid": orcid_val,
                    "affiliation": "; ".join(affs),
                })
        doi = ""
        for aid in art.findall(".//ArticleId"):
            if aid.attrib.get("IdType") == "doi":
                doi = aid.text or ""
                break
        out.append({
            "pmid": pmid, "title": title, "abstract": abstract,
            "authors": "; ".join(authors), "journal": journal,
            "pub_date": pub_date, "iso_date": iso_date, "doi": doi,
            "author_details": author_details,
        })
    # Sort newest first
    out.sort(key=lambda p: (p.get("iso_date") or "", p.get("pmid") or ""), reverse=True)
    return out


# ════════════════════════════════════════════════════════════════════
# NIH Reporter lookup
# ════════════════════════════════════════════════════════════════════
def fetch_nih_funding_by_pi_id(pi_id: int, limit: int = 100):
    """Live query NIH Reporter API for projects by PI profile ID."""
    payload = {
        "criteria": {"pi_profile_ids": [int(pi_id)]},
        "include_fields": [
            "ApplId", "FiscalYear", "ProjectNum", "ProjectTitle",
            "ProjectStartDate", "ProjectEndDate", "AwardAmount",
            "AgencyIcAdmin", "Organization", "PrincipalInvestigators",
            "AbstractText", "Terms",
        ],
        "offset": 0,
        "limit": limit,
        "sort_field": "fiscal_year",
        "sort_order": "desc",
    }
    r = requests.post(NIH_API, json=payload, timeout=60)
    r.raise_for_status()
    data = r.json()
    results = data.get("results") or []
    total = (data.get("meta") or {}).get("total", 0)

    projects = []
    total_funding = 0
    pi_name_from_api = ""
    for p in results:
        amount = p.get("award_amount") or 0
        try:
            total_funding += int(amount)
        except (ValueError, TypeError):
            pass
        org = (p.get("organization") or {}).get("org_name", "") or ""
        agency = (p.get("agency_ic_admin") or {}).get("name", "") or ""

        # Pull PI name from the first matching record's principal_investigators
        if not pi_name_from_api:
            for pi in (p.get("principal_investigators") or []):
                if isinstance(pi, dict) and (pi.get("profile_id") == pi_id or
                                             str(pi.get("profile_id") or "") == str(pi_id)):
                    pi_name_from_api = pi.get("full_name") or ""
                    if pi_name_from_api:
                        break
            # Fall back to first contact PI if no exact match found
            if not pi_name_from_api:
                for pi in (p.get("principal_investigators") or []):
                    if isinstance(pi, dict) and pi.get("is_contact_pi"):
                        pi_name_from_api = pi.get("full_name") or ""
                        if pi_name_from_api:
                            break

        projects.append({
            "appl_id": p.get("appl_id"),
            "fiscal_year": p.get("fiscal_year"),
            "project_num": p.get("project_num"),
            "title": p.get("project_title"),
            "start_date": p.get("project_start_date"),
            "end_date": p.get("project_end_date"),
            "award_amount": amount,
            "agency": agency,
            "organization": org,
        })

    return {
        "pi_id": pi_id,
        "pi_name_from_api": pi_name_from_api,
        "total_projects": total,
        "total_funding": total_funding,
        "projects": projects,
    }


# ════════════════════════════════════════════════════════════════════
# Routes
# ════════════════════════════════════════════════════════════════════
@app.route("/api/track-click", methods=["POST"])
def track_click():
    """Record a user click event."""
    data = request.get_json() or {}

    element = data.get("element", "unknown")
    page = data.get("page", "unknown")
    user_agent = request.headers.get("User-Agent", "")
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(CLICK_DB_PATH)
    c = conn.cursor()
    c.execute("""
        INSERT INTO click_events (element, page, timestamp, user_agent)
        VALUES (?, ?, ?, ?)
    """, (element, page, timestamp, user_agent))
    conn.commit()
    conn.close()

    return jsonify({"status": "success"})


@app.route("/api/click-stats")
def click_stats():
    """Return click counts grouped by clicked element."""
    conn = sqlite3.connect(CLICK_DB_PATH)
    c = conn.cursor()
    c.execute("""
        SELECT element, COUNT(*) AS count
        FROM click_events
        GROUP BY element
        ORDER BY count DESC
    """)
    rows = c.fetchall()
    conn.close()

    stats = [{"element": row[0], "count": row[1]} for row in rows]
    return jsonify(stats)
  
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/members")
def api_members():
    """Return member list for autocomplete."""
    q = request.args.get("q", "").strip()
    if q:
        members = find_members_by_query(q, limit=20)
    else:
        members = load_members()
    return jsonify(members)


@app.route("/api/stats")
def api_stats():
    """Quick database stats for the dashboard."""
    conn = get_db()
    pi_n = conn.execute("SELECT COUNT(*) FROM pi").fetchone()[0]
    proj_n = conn.execute("SELECT COUNT(*) FROM project").fetchone()[0]
    pub_n = conn.execute("SELECT COUNT(*) FROM publication").fetchone()[0]
    conn.close()

    members = load_members()
    members_with_orcid = sum(1 for m in members if m["orcid"])
    members_with_pi_id = sum(1 for m in members if m["pi_id"])

    # Total publications across all MCC members (from Sheet2 PUB_COUNT)
    total_pubs = 0
    for m in members:
        try:
            n = m.get("pub_count")
            if n is not None and n != "":
                total_pubs += int(float(n))
        except (ValueError, TypeError):
            continue

    return jsonify({
        "members_total": len(members),
        "members_with_orcid": members_with_orcid,
        "members_with_pi_id": members_with_pi_id,
        "projects_indexed": proj_n,
        "publications_total": total_pubs,
        "publications_cached": pub_n,
        "pis_in_db": pi_n,
    })


@app.route("/api/mcc_names")
def api_mcc_names():
    """Return all MCC member names for author highlighting."""
    members = load_members()
    names = []
    for m in members:
        raw = m["name"]
        names.append(raw)
        if "," in raw:
            last, first = [s.strip() for s in raw.split(",", 1)]
            first_parts = first.split()
            first_short = first_parts[0] if first_parts else ""
            if first_short:
                names.append(f"{first_short} {last}")
                names.append(f"{first} {last}")
            # Add hyphenated last-name parts as separate entries
            if "-" in last:
                for part in last.split("-"):
                    part = part.strip()
                    if part and first_short:
                        names.append(f"{first_short} {part}")
    return jsonify(sorted(set(names)))


@app.route("/search")
def search_keyword():
    """Live PubMed keyword search scoped to MCC/Weill Cornell affiliations."""
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"publications": [], "total": 0})

    year_start = request.args.get("year_start", "").strip()
    year_end = request.args.get("year_end", "").strip()
    researcher = request.args.get("researcher", "").strip()

    try:
        pubs, total = pubmed_search_by_keyword(
            q, year_start=year_start, year_end=year_end,
            retmax=100,
        )
    except Exception as e:
        return jsonify({"error": f"PubMed search failed: {e}",
                        "publications": [], "total": 0})

    # Keep only publications with at least one MCC member as author
    pubs = [p for p in pubs if _pub_has_mcc_author(p)]

    # Apply researcher filter with ORCID + affiliation verification.
    if researcher:
        r_text = researcher.strip()
        # Determine if this looks like a specific name (has comma or
        # multiple words) vs a single partial term like "todd" or "neal"
        is_specific = "," in r_text or len(r_text.split()) >= 2
        member = find_member_by_name(r_text) if is_specific else None

        if member and member.get("orcid"):
            confirmed = _get_orcid_confirmed_pmids(member["orcid"])
        else:
            confirmed = set()

        filtered = []
        for p in pubs:
            pmid = p.get("pmid", "")
            # ORCID-confirmed PMID — definitely this member's paper
            if confirmed and pmid in confirmed:
                filtered.append(p)
                continue
            # Use name + affiliation matching
            if _pub_matches_researcher(p, r_text, member):
                filtered.append(p)
        pubs = filtered

    # Strip author_details from response (internal use only)
    for p in pubs:
        p.pop("author_details", None)

    return jsonify({"publications": pubs[:50], "total": total})


@app.route("/search_name")
def search_name():
    """Look up a researcher by name → return profile + publications."""
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"error": "Please provide a name."})

    member = find_member_by_name(name)
    if not member:
        return jsonify({"error": f"No MCC member found matching '{name}'."})

    publications = []
    source = "none"

    # Strategy 1: ORCID-based live lookup (best quality)
    if member["orcid"]:
        try:
            result = lookup_publications(member["orcid"], save_to_db=True)
            publications = result.get("publications", [])
            source = "orcid"
        except Exception as e:
            print(f"[WARN] ORCID lookup failed for {name}: {e}")

    # Strategy 2: Cached database lookup (by ORCID)
    if not publications and member["orcid"]:
        conn = get_db()
        rows = conn.execute("""
            SELECT pmid, title, abstract, authors, journal, pub_date, doi
            FROM publication WHERE orcid = ?
            ORDER BY pub_date DESC LIMIT 50
        """, (member["orcid"],)).fetchall()
        conn.close()
        publications = [dict(r) for r in rows]
        source = "cache"

    # Strategy 3: Live PubMed by name (fallback when no ORCID)
    if not publications:
        try:
            publications = pubmed_search_by_name(member["name"], retmax=20)
            source = "pubmed_name"
        except Exception as e:
            print(f"[WARN] Name search failed: {e}")

    return jsonify({
        "name": member["name"],
        "orcid": member["orcid"],
        "pi_id": member["pi_id"],
        "program": member.get("program", ""),
        "vivo_url": member.get("vivo_url", ""),
        "publication_count": len(publications),
        "source": source,
        "publications": publications,
    })


@app.route("/search_funding")
def search_funding():
    """Look up NIH funding by PI_ID or by member name."""
    pi_id = request.args.get("pi_id", "").strip()
    name = request.args.get("name", "").strip()

    member = None
    if name:
        member = find_member_by_name(name)
        if not member:
            return jsonify({"error": f"No MCC member found matching '{name}'."})
        if not member["pi_id"]:
            return jsonify({
                "error": f"{member['name']} has no NIH PI_ID — no NIH funding records to look up.",
                "name": member["name"],
            })
        pi_id = member["pi_id"]

    if not pi_id:
        return jsonify({"error": "Provide either pi_id or name."})

    try:
        pi_id_int = int(float(pi_id))
    except (ValueError, TypeError):
        return jsonify({"error": f"Invalid PI_ID: {pi_id}"})

    try:
        funding = fetch_nih_funding_by_pi_id(pi_id_int, limit=100)
    except Exception as e:
        return jsonify({"error": f"NIH Reporter API error: {e}"})

    # Resolve a researcher name in this priority order:
    #   1. The member typed by the user (already validated above)
    #   2. Local MCC roster lookup by NIH_ID
    #   3. PI name returned in the NIH RePORTER project data itself
    #   4. Empty (frontend will fall back to "NIH_ID: ...")
    resolved_name = ""
    resolved_orcid = ""
    if member:
        resolved_name = member["name"]
        resolved_orcid = member.get("orcid", "")
    else:
        roster_match = find_member_by_pi_id(pi_id_int)
        if roster_match:
            resolved_name = roster_match["name"]
            resolved_orcid = roster_match.get("orcid", "")
        else:
            resolved_name = funding.get("pi_name_from_api", "")

    funding["name"] = resolved_name
    funding["orcid"] = resolved_orcid
    return jsonify(funding)


@app.route("/researcher")
def researcher_combined():
    """Combined view: publications + NIH funding for one or more matching members."""
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"error": "Please provide a name."})

    matches = find_members_fuzzy(name, limit=10)
    if not matches:
        return jsonify({"error": f"No MCC member found matching '{name}'."})

    # If only one match, return the detailed profile directly
    # If multiple, return a list of candidates and load the first one's detail
    results = []
    for match_info in matches:
        member = match_info["member"]
        match_type = match_info["match_type"]
        entry = {
            "name": member["name"],
            "orcid": member["orcid"],
            "pi_id": member["pi_id"],
            "program": member.get("program", ""),
            "vivo_url": member.get("vivo_url", ""),
            "match_type": match_type,
        }
        results.append(entry)

    # Load full detail for the first (best) match
    member = matches[0]["member"]
    response = {
        "name": member["name"],
        "orcid": member["orcid"],
        "pi_id": member["pi_id"],
        "program": member.get("program", ""),
        "vivo_url": member.get("vivo_url", ""),
        "match_type": matches[0]["match_type"],
        "all_matches": results,
        "publications": [],
        "publication_count": 0,
        "publication_source": "none",
        "funding": None,
        "funding_error": None,
    }

    # Publications
    if member["orcid"]:
        try:
            r = lookup_publications(member["orcid"], save_to_db=True)
            response["publications"] = r.get("publications", [])
            response["publication_count"] = len(response["publications"])
            response["publication_source"] = "orcid"
        except Exception as e:
            response["publication_source"] = f"orcid_error: {e}"

    if not response["publications"]:
        try:
            response["publications"] = pubmed_search_by_name(member["name"], retmax=15)
            response["publication_count"] = len(response["publications"])
            response["publication_source"] = "pubmed_name"
        except Exception as e:
            response["publication_source"] = f"name_error: {e}"

    # Funding
    if member["pi_id"]:
        try:
            response["funding"] = fetch_nih_funding_by_pi_id(int(float(member["pi_id"])), limit=50)
        except Exception as e:
            response["funding_error"] = str(e)
    else:
        response["funding_error"] = "No NIH PI_ID on file for this member."

    return jsonify(response)


def _parse_nih_keyword(raw: str) -> tuple:
    """Parse a funding keyword query, respecting quoted exact phrases.

    Returns (search_text, operator, exact_phrases):
      - Quoted phrase like "vitamin d" → exact phrase, operator "and"
      - Unquoted terms → fuzzy-expanded, operator "or"
      exact_phrases is a list of lowered phrases the user explicitly quoted.
    """
    import re
    raw = raw.strip()
    # Check for quoted exact phrase: "vitamin d" or 'breast cancer'
    m = re.match(r'^["\'](.+?)["\']$', raw)
    if m:
        phrase = m.group(1).strip()
        # For hyphenated terms like GLP-1, also add parts without hyphen
        # so the NIH API can actually find something (it ignores "GLP-1")
        search_parts = [phrase]
        if "-" in phrase:
            search_parts.extend(p for p in phrase.split("-") if len(p) >= 2)
        return " ".join(search_parts), "and", [phrase.lower()]
    # Unquoted → expand for fuzzy matching; also add de-hyphenated parts
    expanded = _expand_query_for_nih(raw)
    for w in raw.split():
        if "-" in w:
            for part in w.split("-"):
                if len(part) >= 2:
                    expanded += " " + part
    return expanded, "or", []


def _nih_project_matches_keyword(project: dict, keywords: list,
                                  exact_phrases: list) -> bool:
    """Server-side check that a project actually contains the keyword(s).

    The NIH Reporter API sometimes ignores text filters for certain terms
    (especially hyphenated like GLP-1). This post-filter catches those cases.
    """
    # Build searchable text from title, terms, and abstract
    parts = [
        project.get("project_title") or "",
        project.get("terms") or "",
        project.get("abstract_text") or "",
    ]
    haystack = " ".join(parts).lower()

    # If user quoted an exact phrase, require it
    if exact_phrases:
        return all(phrase in haystack for phrase in exact_phrases)

    # For unquoted search: require at least one keyword to appear
    if keywords:
        return any(kw in haystack for kw in keywords)

    return True


@app.route("/search_funding_keyword")
def search_funding_keyword():
    """Search NIH funding by keyword/topic, with optional PI name filter.

    Query params:
      q  – project keyword / topic (supports quoted exact phrases)
      pi – optional PI name filter (fuzzy-matched against MCC roster)
    """
    q = request.args.get("q", "").strip()
    pi_name = request.args.get("pi", "").strip()

    if not q and not pi_name:
        return jsonify({"error": "Please provide a keyword or PI name."})

    members = load_members()
    matched_names = []

    if pi_name:
        # PI filter: resolve to MCC roster members and use their PI_IDs.
        # This ensures only MCC-member projects appear — never non-MCC PIs.
        matches = find_members_fuzzy(pi_name, limit=5)
        if not matches:
            return jsonify({"error": f"No MCC member found matching '{pi_name}'."})
        pi_ids = []
        for match_info in matches:
            m = match_info["member"]
            if m["pi_id"]:
                pi_ids.append(int(float(m["pi_id"])))
                matched_names.append(m["name"])
        if not pi_ids:
            return jsonify({
                "error": f"{matches[0]['member']['name']} has no NIH PI_ID.",
            })
    else:
        # No PI filter: search all MCC members
        pi_ids = [int(float(m["pi_id"])) for m in members if m["pi_id"]]

    if not pi_ids:
        return jsonify({"error": "No MCC members with NIH PI_IDs found."})

    # Build NIH Reporter query criteria
    criteria = {"pi_profile_ids": pi_ids}

    # Parse keyword, get exact phrases for post-filtering
    exact_phrases = []
    plain_keywords = []
    if q:
        search_text, operator, exact_phrases = _parse_nih_keyword(q)
        criteria["advanced_text_search"] = {
            "operator": operator,
            "search_field": "projecttitle,terms,abstracttext",
            "search_text": search_text,
        }
        # For unquoted queries, extract the original words for post-filter.
        # Include the full term (e.g. "glp-1") plus individual words ≥2 chars.
        if not exact_phrases:
            plain_keywords = [q.lower()]
            plain_keywords.extend(
                w.lower() for w in q.split() if len(w) >= 2 and w.lower() != q.lower()
            )

    # Request extra results so we have enough after post-filtering.
    # NIH API ignores certain terms (hyphenated like GLP-1), so we
    # fetch more and post-filter to find actual matches.
    fetch_limit = 500 if q else 50

    payload = {
        "criteria": criteria,
        "include_fields": [
            "ApplId", "FiscalYear", "ProjectNum", "ProjectTitle",
            "ProjectStartDate", "ProjectEndDate",
            "AgencyIcAdmin", "Organization", "PrincipalInvestigators",
            "AbstractText", "Terms",
        ],
        "offset": 0,
        "limit": fetch_limit,
        "sort_field": "fiscal_year",
        "sort_order": "desc",
    }
    try:
        r = requests.post(NIH_API, json=payload, timeout=60)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        return jsonify({"error": f"NIH Reporter API error: {e}"})

    raw_results = data.get("results") or []

    # Post-filter: verify keyword actually appears in project data
    # (NIH API sometimes ignores text filters for hyphenated terms like GLP-1)
    if q:
        raw_results = [
            p for p in raw_results
            if _nih_project_matches_keyword(p, plain_keywords, exact_phrases)
        ]

    projects = []
    for p in raw_results[:50]:
        org = (p.get("organization") or {}).get("org_name", "") or ""
        agency = (p.get("agency_ic_admin") or {}).get("name", "") or ""
        pis = p.get("principal_investigators") or []
        pi_names = [pi.get("full_name", "") for pi in pis if isinstance(pi, dict) and pi.get("full_name")]
        projects.append({
            "appl_id": p.get("appl_id"),
            "fiscal_year": p.get("fiscal_year"),
            "project_num": p.get("project_num"),
            "title": p.get("project_title"),
            "start_date": p.get("project_start_date"),
            "end_date": p.get("project_end_date"),
            "agency": agency,
            "organization": org,
            "pi_names": "; ".join(pi_names),
        })

    return jsonify({
        "query": q,
        "pi_filter": ", ".join(matched_names) if matched_names else "",
        "total_results": len(projects),
        "projects": projects,
    })


# ════════════════════════════════════════════════════════════════════
# Run
# ════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"[INFO] DB:    {DB_PATH}")
    print(f"[INFO] Excel: {PI_EXCEL}")
    print(f"[INFO] Loaded {len(load_members())} MCC members")
    port = int(os.environ.get("PORT", 5001))
    debug = os.environ.get("FLASK_ENV") != "production"
    app.run(host="0.0.0.0", port=port, debug=debug)
